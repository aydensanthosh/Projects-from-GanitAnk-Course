# importing all the required libraries/modules
import os
from pathlib import Path
import sys
from rich.console import Console
from rich.text import Text
from rich.panel import Panel
from rich.progress import track
from time import sleep
import requests
from openpyxl import Workbook,load_workbook
import json
import datetime
import pyfiglet
from random import randint
import time
from colorama import Fore, init
init(autoreset=True)#Resets the color after every color change.
console=Console()
#activating the excel or creating them if not existing and giving their headers
def Create_folder():
    Path("data").mkdir(exist_ok=True)
Create_folder()
file_headers = {
    "advice":["Advice","ID","Timestamp","Status(Success/Error)"],
    "book":["Book title","Book No.","Timestamp","Status(Success/Error)"],
    "quote":["Quote Type","Quote","Timestamp","Status(Success/Error)"],
    "dog":["Status","Dog pic Link","Timestamp"]
}

#Storing all the data Excels together
DATA_FILES = {
    'advice': 'data/advice_data.xlsx',
    'book': 'data/books_data.xlsx',
    'quote': 'data/quotes_data.xlsx',
    'dog': 'data/dogs_data.xlsx'
}
#Storing all the API links together.
URL = {
    'advice': 'https://api.adviceslip.com/advice',
    'book': 'https://potterapi-fedeperin.vercel.app/en/books',
    'quote': 'https://hindi-quotes.vercel.app/random',
    'dog': 'https://dog.ceo/api/breeds/image/random'
}
def activate_excel():
    """"Creating the files in location if not existing already and initializing them by giving the header names"""
    for file_name , header in file_headers.items():
        wb=Workbook()
        ws=wb.active
        ws.append(header)
        wb.save(DATA_FILES[file_name])

activate_excel()
def API_call(key):
    url=URL[key]
    response=requests.get(url)
    return response.json()

def get_advice():
    """"Fetches the advice from the API and saves it to excel and then displays it on the screen"""
    
    Waiting("advice")
    try:
        current_advice=API_call("advice")["slip"]
        Time=datetime.datetime.now()
        save_advice=[current_advice["advice"],current_advice["id"],Time,"Succesful"]
        console.print(Panel(f"The Advice is: {current_advice["advice"]}",style="Yellow",title=f"ADVICE",subtitle=f"Advice No.{current_advice["id"]}"))
        wb = load_workbook('data/advice_data.xlsx',read_only=False,)
        ws = wb.active
        ws.append(save_advice)
        wb.save('data/advice_data.xlsx')
    except Exception as e:
        console.print(Panel(f"{e}",style="Red",title="ERROR",subtitle="Error 404"))
        
def get_book():
    """"Fetches the Harry potter book recommendation from the API and saves it to excel and then displays it on the screen"""
    try:
        Waiting("book")
        response=API_call("book")
        randomising=randint(1,8)
        Time=datetime.datetime.now()
        for i in response:
            if i["number"]==randomising:
                recomm=i
        console.print(Panel(f"\n{recomm["description"]}\n",style="Yellow",title=f"{recomm["title"]}".upper(),subtitle=f"No. of Pages={recomm["pages"]}"))
        save_book=[recomm["title"],recomm["number"],Time]
        wb = load_workbook(DATA_FILES["book"],read_only=False,) 
        ws = wb.active
        ws.append(save_book)
        wb.save(DATA_FILES["book"])
    except Exception as e:
        console.print(Panel(f"{e}",style="Red",title="ERROR",subtitle="Error 404"))
        
def get_quote():
    """"Fetches the quote from the API and saves it to excel and then displays it on the screen"""
    try:
        Waiting("quote")
        response=requests.get(URL["quote"])
        response=response.json()
        Time=datetime.datetime.now()
        quote=response
        save_quote=[quote["type"],quote["quote"],Time,"Success"]
        wb = load_workbook(DATA_FILES["quote"],read_only=False,) 
        ws = wb.active
        ws.append(save_quote)
        wb.save(DATA_FILES["quote"])
        console.print(Panel(f"\n{quote["quote"]}\n",style="Yellow",title=f"This quote is about {quote["type"].upper()}"))
    except Exception as e:
        console.print(Panel(f"{e}",style="Red",title="ERROR",subtitle="Error 404"))
        
def get_dog():
    """"Fetches the dog image url from the API and saves it to excel and then displays it on the screen"""
    try:
        Waiting("dog")
        response=requests.get(URL["dog"])
        response=response.json()
        Time=datetime.datetime.now()
        dog=response
        save_dog=[dog["status"],dog["message"],Time]
        wb = load_workbook(DATA_FILES["dog"],read_only=False,) 
        ws = wb.active
        ws.append(save_dog)
        wb.save(DATA_FILES["dog"])
        console.print(Panel(f"\n{dog["message"]}\n",style="Yellow",title=f"Dog Picture",subtitle="Copy and paste in your Browser, to view image"))
    except Exception as e:
        console.print(Panel(f"{e}",style="Red",title="ERROR",subtitle="Error 404"))
        
def get_everything():
    """Fetch all content types - daily dose of motivation!"""
    print(f"\n{Fore.CYAN}🌟 Preparing your daily dose of motivation...\n")
    
    print(f"{Fore.YELLOW}{'='*100}")
    get_advice()
    
    print(f"\n{Fore.YELLOW}{'='*100}")
    get_quote()
    
    print(f"\n{Fore.YELLOW}{'='*100}")
    get_book()
    
    print(f"\n{Fore.YELLOW}{'='*100}")
    get_dog()
    
    print(f"\n{Fore.GREEN}🎉 Your daily motivation dose is complete! Have a great day! 🎉")

def show_statistics():
    """Display statistics about saved data"""
    stats = {}
    total_entries = 0
    
    for data_type, file_path in DATA_FILES.items():
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                count = ws.max_row - 1  # Subtract header row
                stats[data_type] = max(0, count)
                total_entries += stats[data_type]
            else:
                stats[data_type] = 0
        except:
            stats[data_type] = 0
    
    stats_info = {
        "📝 Total Entries": total_entries,
        "💡 Advice Count": stats['advice'],
        "📚 Books Count": stats['book'],
        "🎯 Quotes Count": stats['quote'],
        "🐕 Dogs Count": stats['dog']
    }
    console.print(Panel(f"\n{stats_info}\n",title="📊 DATA STATISTICS",style="CYAN"))
    
def process_data():
    sleep(0.02)
    
def Waiting(Process):
    """This is the progress bar displayed when we are waiting for the API to fetch the result"""
    if Process=="advice":
        for i in track(range(100), description='[green]Fetching Advice:'):
            process_data()
    elif Process=="book":
        for i in track(range(100), description='[green]Fetching Book Recommendation:'):
            process_data()
    elif Process=="quote":
        for i in track(range(100), description='[green]Fetching Inspirational quotes:'):
            process_data()
    elif Process=="dog":
        for i in track(range(100), description='[green]Fetching Dog Pictures:'):
            process_data()
    elif Process=="everything":
        pass
    
    
    print(Fore.GREEN + "Here is your request.\n")
def welcome():
    """This is the display of the welcome sign"""
    var="""╔══════════════════════════════════════════════════════════════╗
║                     DAILY MOTIVATION HUB                     ║
║                                                              ║
║         Your one-stop destination for daily inspiration!     ║
║                                                              ║
║   Books  |   Advice  |   Quotes  |   Cute Dogs               ║
╚══════════════════════════════════════════════════════════════╝\n"""
    print(Fore.GREEN + var)
def menu():
    """This is the display of what functions this program can do."""
    var1="""┌────────────────────MAIN─MENU───────────────────┐
│                                                │
│                                                │
│  1. Get Random Advice                          │
│                                                │     
│  2. Get Harry Potter Book Recommendation       │
│                                                │
│  3. Get Inspirational Quotes                   │
│                                                │
│  4. Get Cute Dog Pics (Maybe)                  │
│                                                │
│  5. Get One of Each of the above               │
│                                                │
│  6. Show Data Statistics                       │
│                                                │
│  7. Exit                                       │
└────────────────────────────────────────────────┘\n"""
    print(Fore.CYAN + var1)

###########################################################################################################################
#Main APP loop
def main():
    # Check Python version
    if sys.version_info < (3, 12):
        print("Error: This application requires Python 3.12 or higher")
        sys.exit(1)
    
    welcome()
    print("\n\n")
    while True:
        menu()
        choice=input(Fore.CYAN + "Enter your choice (1-7):")
        if choice=='1':
            get_advice()
        elif choice=='2':
            get_book()
        elif choice=='3':
            get_quote()
        elif choice=='4':
            get_dog()
        elif choice=='5':
            get_everything()        
        elif choice=='6':
            show_statistics()
        elif choice=='7':
            print(f"\n{Fore.GREEN}Thank you for using Daily Motivation Hub! 🌟")
            print(f"Stay motivated and have a great day! 💪")
            break
        else:
            t = Text("Warning: Invalid Choice")
            t.stylize("bold red")
            t.append(" — Only 1-7 Allowed!", style="italic yellow")
            console.print(t)
                #print(f"{Fore.RED}❌ Invalid choice. Please enter a number between 1-7.")
        if choice in ['1','2','3','4','5','6']:
            input(Fore.BLACK + "Press Enter to Continue....\t\t\t")
main()    


""""This program will provide you with advice, Harry potter book recommendation, Inspirational quotes and cute dog pics
    It will save all the advices, recommendations, quotes and pic links to their respective Excel sheets (data\\advice_data.xlsx)
    and will display"""